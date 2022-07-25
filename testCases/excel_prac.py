import openpyxl

path = "C:/Users/sbhutkar/PycharmProjects/DemoWebShop/testData/testData.xlsx"

workbook = openpyxl.load_workbook(path)
sheet = workbook.get_sheet_by_name("Sheet1")
print(sheet.max_row)
print(sheet.max_column)
print(sheet.cell(row=2,column=2).value)
sheet.cell(row=2,column=2).value = "sandy"
workbook.save(path)



